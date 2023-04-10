from binance.client import Client
import datetime
import pandas as pd
import openpyxl as op
from openpyxl.styles import numbers
import time

with open('ключи.txt', 'r') as f:
    api = f.readline().strip()
    secret = f.readline().strip()


api = api.split('=')[1].strip()
secret = secret.split('=')[1].strip()
client = Client(api, secret)
allpositions = []
alltickers = ['BTCUSDT','ADAUSDT','BATUSDT','BNBUSDT','EOSUSDT','ETCUSDT','ETHUSDT','HOTUSDT',
                'ICXUSDT','IOTAUSDT','LINKUSDT','LTCUSDT','NEOUSDT','ONTUSDT','QTUMUSDT','TRXUSDT',
                'VETUSDT','WAVESUSDT','XLMUSDT','XMRUSDT','XRPUSDT','ZILUSDT','ZRXUSDT','ZECUSDT']
    
for ticker in alltickers:
    closed_positions = client.futures_account_trades(symbol=ticker)
    allpositions.append(closed_positions)
    print(f'Собираю: {ticker}')

workbook = op.load_workbook('Журнал сделок.xlsx')
sheet = workbook.get_sheet_by_name('Лист1')
df = pd.DataFrame(sheet.values)
for closed_positions in allpositions:
    if closed_positions:
        opencheck = False
        ordernumber = 0
        open_price = 0
        close_price = 0
        commission = 0
        oprice = 0
        cprice = 0
        pnl = 0
        nopenorder = 0
        ncloseorder = 0
        qty = 0
        volumeo = 0
        volumec = 0
        firstcheck = False
        
        for position in closed_positions:
            if float(position.get('realizedPnl')) == 0:
                if ordernumber != position.get('orderId'):
                    
                    open_price = 0
                    close_price = 0
                    commission = 0
                    oprice = 0
                    cprice = 0
                    pnl = 0
                    nopenorder = 0
                    ncloseorder = 0
                    qty = 0
                    volumeo = 0
                    volumec = 0
                opencheck = True
                ordernumber = position.get('orderId')
                
            if opencheck and ordernumber == position.get('orderId'):
                oprice += float(position.get('price'))
                nopenorder += 1
                open_price = round(oprice/nopenorder,len(str(float(position.get('price'))).split('.')[1]))
                commission += float(position.get('commission'))
                pnl += float(position.get('realizedPnl'))
                qty += float(position.get('quoteQty'))
                timestamp = position.get('time') / 1000 
                dt_object = datetime.datetime.fromtimestamp(timestamp)
                open_date = dt_object.strftime("%d-%m-%Y")
                open_time = dt_object.strftime("%H:%M:%S")
                symbol = position.get('symbol')
                side = position.get('side')
                volumeo += float(position.get('qty'))
                
            if opencheck and ordernumber != position.get('orderId'):
                cprice += float(position.get('price'))
                ncloseorder += 1
                close_price = round(cprice/ncloseorder,len(str(float(position.get('price'))).split('.')[1]))
                commission += float(position.get('commission'))
                pnl += float(position.get('realizedPnl'))
                timestamp = position.get('time') / 1000 
                dt_object = datetime.datetime.fromtimestamp(timestamp)
                close_date = dt_object.strftime("%d-%m-%Y")
                close_time = dt_object.strftime("%H:%M:%S")
                
            if opencheck and float(position.get('realizedPnl')) != 0:
                volumec += float(position.get('qty'))
                if abs(volumec - volumeo) / max(abs(volumec), abs(volumeo)) < 0.001:
                    procentchange = round((close_price-open_price)/open_price*100,5)
                    values = list(df[8].unique())
                    if close_price in values:
                        continue
                    for i in range(1,len(df)):
                        if df.loc[i,1] == None:
                            df.loc[i,0] = i
                            df.loc[i,1] = open_date
                            df.loc[i,2] = open_time
                            df.loc[i,3] = close_date
                            df.loc[i,4] = close_time
                            df.loc[i,5] = symbol
                            df.loc[i,6] = side
                            df.loc[i,7] = open_price
                            df.loc[i,8] = close_price
                            df.loc[i,9] = qty
                            df.loc[i,10] = commission
                            df.loc[i,11] = pnl
                            if side == 'BUY':
                                df.loc[i,12] = procentchange
                            if side == 'SELL':
                                df.loc[i,12] = -procentchange
                            break

for i in range(1,len(df)):
    if df.loc[i,1] == None:
        break
    df.loc[i,3] = int(datetime.datetime.strptime(df.loc[i,3], '%d-%m-%Y').timestamp())
df[1:] = df[1:].sort_values(3)
for i in range(1,len(df)):
    if df.loc[i,1] == None:
        break
    timestamp = df.loc[i,3]
    dt_object = datetime.datetime.fromtimestamp(timestamp)
    df.loc[i,3] = dt_object.strftime("%d-%m-%Y")



for i in range(2, len(df)):
    sheet.cell(row=i, column=1).value = i-1
    sheet.cell(row=i, column=2).value = df.loc[i,1]
    sheet.cell(row=i, column=3).value = df.loc[i,2]
    sheet.cell(row=i, column=4).value = df.loc[i,3]
    sheet.cell(row=i, column=5).value = df.loc[i,4]
    sheet.cell(row=i, column=6).value = df.loc[i,5]
    sheet.cell(row=i, column=7).value = df.loc[i,6]
    sheet.cell(row=i, column=8).value = df.loc[i,7]
    sheet.cell(row=i, column=9).value = df.loc[i,8]
    sheet.cell(row=i, column=10).value = df.loc[i,9]
    sheet.cell(row=i, column=11).value = df.loc[i,10]
    sheet.cell(row=i, column=12).value = df.loc[i,11]
    sheet.cell(row=i, column=13).value = df.loc[i,12]
# df = pd.DataFrame(sheet.values)
# df.head(30)
workbook.save("Журнал сделок.xlsx")