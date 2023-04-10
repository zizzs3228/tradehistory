import sys
from PyQt5.QtWidgets import QApplication, QWidget, QLabel, QDateEdit, QLineEdit, QPushButton, QVBoxLayout, QHBoxLayout, QMessageBox, QGroupBox, QGridLayout
from PyQt5.QtGui import QColor, QPalette
from PyQt5.QtCore import QDate
from PyQt5.QtCore import Qt
import datetime
from binance.client import Client
import pandas as pd
import openpyxl as op
import time

class DateSelector(QWidget):
    def __init__(self):
        super().__init__()

        # Set background color for the window
        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(61, 61, 61))
        self.setPalette(palette)

        # Create a label and two date selectors
        self.date_label = QLabel("Выбери дату начала сбора и дату окончания сбора:")
        self.date_label.setAlignment(Qt.AlignCenter)
        self.date1 = QDateEdit()
        self.date2 = QDateEdit()

        # Center the date_label text
        self.date_label.setStyleSheet("QLabel { font-size: 20px; color: #ffffff; }")

        # Create a label and a text field
        self.text_label = QLabel("Введи токены, например TRXUSDT, BNBUSDT:")
        self.text_field = QLineEdit()

        # Create a submit button
        self.submit_button = QPushButton("СБОР")
        self.submit_button.clicked.connect(self.submit)

        # Set the font size and color for the labels and button
        font = self.font()
        font.setPointSize(16)
        self.date_label.setFont(font)
        self.date_label.setStyleSheet("QLabel { font-size: 20px; color: #ffffff; }")
        self.text_label.setFont(font)
        self.text_label.setStyleSheet("QLabel { font-size: 20px; color: #ffffff; }")
        self.submit_button.setFont(font)
        self.submit_button.setStyleSheet("QPushButton { color: #ffffff; background-color: #c62828; }")

        # Set the size of the date selectors
        self.date2.setDate(QDate.currentDate())

        # Calculate the date one month ago from today's date
        date_one_month_ago = QDate.currentDate().addMonths(-1)
        self.date1.setDate(date_one_month_ago)
        self.date1.setDisplayFormat("dd-MM-yyyy")
        self.date2.setDisplayFormat("dd-MM-yyyy")
        self.date1.setFixedSize(150, 40)
        self.date2.setFixedSize(150, 40)

        # Set the size of the text field
        self.text_field.setFixedSize(300, 40)
        self.text_field.setStyleSheet("QLineEdit { color: #ffffff; background-color: #424242; }")

        # Create a grid layout for the labels and date selectors
        date_layout = QGridLayout()
        date_layout.addWidget(self.date_label, 0, 0, 1, 2)
        date_layout.addWidget(self.date1, 1, 0)
        date_layout.addWidget(self.date2, 1, 1)

        # Create a horizontal layout for the text label and field
        text_layout = QHBoxLayout()
        text_layout.addWidget(self.text_label)
        text_layout.addWidget(self.text_field)

        # Create a group box for the date selectors
        date_box = QGroupBox("Date Selector")
        date_box.setAlignment(Qt.AlignCenter)
        date_box.setLayout(date_layout)
        date_box.setStyleSheet("QGroupBox { color: #ffffff; background-color: #303030; }")

        # Create a group box for the text field
        text_box = QGroupBox("Text Field")
        text_box.setAlignment(Qt.AlignCenter)
        text_box.setLayout(text_layout)
        text_box.setStyleSheet("QGroupBox { color: #ffffff; background-color: #303030; }")

        # Create a vertical layout for the group boxes and submit button
        layout = QVBoxLayout()
        layout.addWidget(date_box)
        layout.addWidget(text_box)
        layout.addWidget(self.submit_button)

        # Set the layout for the main widget
        self.setLayout(layout)

    def submit(self):
        # Get the values from the widgets
        self.date1_value = self.date1.date().toString("dd-MM-yyyy")
        self.date2_value = self.date2.date().toString("dd-MM-yyyy")
        self.text_value = self.text_field.text()

        self.date_start = datetime.datetime.strptime(self.date1_value, "%d-%m-%Y")
        self.date_end = datetime.datetime.strptime(self.date2_value, "%d-%m-%Y")

        self.alltickers = ['BTCUSDT','ADAUSDT','BATUSDT','BNBUSDT','EOSUSDT','ETCUSDT','ETHUSDT','HOTUSDT',
                      'ICXUSDT','IOTAUSDT','LINKUSDT','LTCUSDT','NEOUSDT','ONTUSDT','QTUMUSDT','TRXUSDT',
                      'VETUSDT','WAVESUSDT','XLMUSDT','XMRUSDT','XRPUSDT','ZILUSDT','ZRXUSDT','ZECUSDT']
        
        if self.text_value:
            self.alltickers = self.text_value.split(',')
            self.alltickers = [x.strip() for x in self.alltickers]
        # Get the Unix timestamp from the datetime object
        # self.date_start = int(self.date_start.timestamp())*1000
        # self.date_end = int(self.date_end.timestamp())*1000
        
        # Do something with the values (e.g., print them to the console)
        print("Дата начала сбора:", self.date1_value)
        print("Дата окончания сбора:", self.date2_value)
        print("Тикеры, которые будут собраны:", self.alltickers)
        
        self.close()
        self.orderstory()
    
    def orderstory(self):
        with open('ключи.txt', 'r') as f:
            api = f.readline().strip()
            secret = f.readline().strip()


        api = api.split('=')[1].strip()
        secret = secret.split('=')[1].strip()
        client = Client(api, secret)

        
        
        allpositions = []
        date_string1 = self.date1_value
        date_string2 = self.date2_value
        start_time = int(datetime.datetime.strptime(date_string1, '%d-%m-%Y').timestamp())*1000
        end_time = int(datetime.datetime.strptime(date_string2, '%d-%m-%Y').timestamp())*1000
        firstdate = start_time
        datepairs = []
        while True:
            if firstdate+(7*24*60*60*1000)>end_time:
                datepairs.append((firstdate,end_time))
                break
            datepairs.append((firstdate,firstdate+(7*24*60*60*1000)))
            firstdate = firstdate+(3*24*60*60*1000)
            
        for ticker in self.alltickers:
            for pair in datepairs:
                closed_positions = client.futures_account_trades(symbol=ticker,startTime=pair[0],endTime=pair[1])
                allpositions.append(closed_positions)
            print(f'Собираю {ticker}')
            time.sleep(1)
        # closed_positions = client.futures_account_trades(startTime=start_time, endTime=end_time)
        # for pair in datepairs:
        #     timestamp_ms = 1645407600000
        #     date_string1 = datetime.datetime.fromtimestamp(pair[0] / 1000.0).strftime('%Y-%m-%d')
        #     date_string2 = datetime.datetime.fromtimestamp(pair[1] / 1000.0).strftime('%Y-%m-%d')
        #     print(date_string1,date_string2)
        workbook = op.load_workbook('shablon.xlsx')
        sheet = workbook.get_sheet_by_name('Лист1')
        df = pd.DataFrame(sheet.values)
        for closed_positions in allpositions:
            if closed_positions:
                opencheck = False
                ordernumber = 0
                open_price = 0
                close_price = 0
                commission = 0
                oprice = 0
                cprice = 0
                pnl = 0
                nopenorder = 0
                ncloseorder = 0
                qty = 0
                volumeo = 0
                volumec = 0
                firstcheck = False
                
                for position in closed_positions:
                    if float(position.get('realizedPnl')) == 0:
                        if ordernumber != position.get('orderId'):
                            
                            open_price = 0
                            close_price = 0
                            commission = 0
                            oprice = 0
                            cprice = 0
                            pnl = 0
                            nopenorder = 0
                            ncloseorder = 0
                            qty = 0
                            volumeo = 0
                            volumec = 0
                        opencheck = True
                        ordernumber = position.get('orderId')
                        
                    if opencheck and ordernumber == position.get('orderId'):
                        oprice += float(position.get('price'))
                        nopenorder += 1
                        open_price = round(oprice/nopenorder,len(str(float(position.get('price'))).split('.')[1]))
                        commission += float(position.get('commission'))
                        pnl += float(position.get('realizedPnl'))
                        qty += float(position.get('quoteQty'))
                        timestamp = position.get('time') / 1000 
                        dt_object = datetime.datetime.fromtimestamp(timestamp)
                        open_date = dt_object.strftime("%d-%m-%Y")
                        open_time = dt_object.strftime("%H:%M:%S")
                        symbol = position.get('symbol')
                        side = position.get('side')
                        volumeo += float(position.get('qty'))
                        
                    if opencheck and ordernumber != position.get('orderId'):
                        cprice += float(position.get('price'))
                        ncloseorder += 1
                        close_price = round(cprice/ncloseorder,len(str(float(position.get('price'))).split('.')[1]))
                        commission += float(position.get('commission'))
                        pnl += float(position.get('realizedPnl'))
                        timestamp = position.get('time') / 1000 
                        dt_object = datetime.datetime.fromtimestamp(timestamp)
                        close_date = dt_object.strftime("%d-%m-%Y")
                        close_time = dt_object.strftime("%H:%M:%S")
                        
                    if opencheck and float(position.get('realizedPnl')) != 0:
                        volumec += float(position.get('qty'))
                        if abs(volumec - volumeo) / max(abs(volumec), abs(volumeo)) < 0.001:
                            procentchange = round((close_price-open_price)/open_price*100,5)
                            values = list(df[8].unique())
                            if close_price in values:
                                continue
                            for i in range(1,len(df)):
                                if df.loc[i,1] == None:
                                    df.loc[i,0] = i
                                    df.loc[i,1] = open_date
                                    df.loc[i,2] = open_time
                                    df.loc[i,3] = close_date
                                    df.loc[i,4] = close_time
                                    df.loc[i,5] = symbol
                                    df.loc[i,6] = side
                                    df.loc[i,7] = open_price
                                    df.loc[i,8] = close_price
                                    df.loc[i,9] = qty
                                    df.loc[i,10] = commission
                                    df.loc[i,11] = pnl
                                    if side == 'BUY':
                                        df.loc[i,12] = procentchange
                                    if side == 'SELL':
                                        df.loc[i,12] = -procentchange
                                    break

        for i in range(1,len(df)):
            if df.loc[i,1] == None:
                break
            df.loc[i,3] = int(datetime.datetime.strptime(df.loc[i,3], '%d-%m-%Y').timestamp())
        df[1:] = df[1:].sort_values(3)
        for i in range(1,len(df)):
            if df.loc[i,1] == None:
                break
            timestamp = df.loc[i,3]
            dt_object = datetime.datetime.fromtimestamp(timestamp)
            df.loc[i,3] = dt_object.strftime("%d-%m-%Y")



        for i in range(2, len(df)):
            sheet.cell(row=i, column=1).value = i-1
            sheet.cell(row=i, column=2).value = df.loc[i,1]
            sheet.cell(row=i, column=3).value = df.loc[i,2]
            sheet.cell(row=i, column=4).value = df.loc[i,3]
            sheet.cell(row=i, column=5).value = df.loc[i,4]
            sheet.cell(row=i, column=6).value = df.loc[i,5]
            sheet.cell(row=i, column=7).value = df.loc[i,6]
            sheet.cell(row=i, column=8).value = df.loc[i,7]
            sheet.cell(row=i, column=9).value = df.loc[i,8]
            sheet.cell(row=i, column=10).value = df.loc[i,9]
            sheet.cell(row=i, column=11).value = df.loc[i,10]
            sheet.cell(row=i, column=12).value = df.loc[i,11]
            sheet.cell(row=i, column=13).value = df.loc[i,12]
        # df = pd.DataFrame(sheet.values)
        # df.head(30)
        workbook.save(f"{self.date1_value}-{self.date2_value}.xlsx")
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = DateSelector()
    window.show()
    sys.exit(app.exec_())
